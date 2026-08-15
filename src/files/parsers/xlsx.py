"""Parser spreadsheet dengan batas memori eksplisit untuk XLSX dan CSV."""

import csv
from typing import Any

from src.core.config import settings


class SpreadsheetParser:
    @staticmethod
    def parse_xlsx(file_path: str) -> tuple[str | None, dict[str, Any] | None]:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            structured_data: dict[str, Any] = {}
            text_lines: list[str] = []
            total_rows = 0
            total_cells = 0
            total_chars = 0
            truncated = False
            try:
                for sheetname in wb.sheetnames:
                    sheet = wb[sheetname]
                    rows_data: list[list[str]] = []
                    for row in sheet.iter_rows(values_only=True):
                        if not any(cell is not None for cell in row):
                            continue
                        str_row = [
                            (str(cell) if cell is not None else "")[:10000]
                            for cell in row
                        ]
                        next_cells = total_cells + len(str_row)
                        row_chars = sum(len(cell) for cell in str_row)
                        if (
                            total_rows >= settings.max_spreadsheet_rows
                            or next_cells > settings.max_spreadsheet_cells
                            or total_chars + row_chars > settings.max_document_extract_chars
                        ):
                            truncated = True
                            break
                        rows_data.append(str_row)
                        total_rows += 1
                        total_cells = next_cells
                        total_chars += row_chars
                        if len(text_lines) < 100:
                            text_lines.append(" | ".join(str_row))
                    structured_data[sheetname] = rows_data
                    if truncated:
                        break
            finally:
                wb.close()

            suffix = "\n[Data dipotong karena batas keamanan parser.]" if truncated else ""
            summary_text = (
                f"=== Spreadsheet XLSX (Sheets: {len(structured_data)}, Rows: {total_rows}) ===\n"
                + "\n".join(text_lines)
                + suffix
            )
            return summary_text, structured_data
        except Exception as exc:
            return f"Error parsing XLSX: {exc}", None

    @staticmethod
    def parse_csv(file_path: str) -> tuple[str | None, dict[str, Any] | None]:
        try:
            rows: list[list[str]] = []
            total_cells = 0
            total_chars = 0
            truncated = False
            with open(file_path, "r", encoding="utf-8", errors="replace", newline="") as handle:
                reader = csv.reader(handle)
                for row in reader:
                    if not row:
                        continue
                    bounded_row = [cell[:10000] for cell in row]
                    row_chars = sum(len(cell) for cell in bounded_row)
                    if (
                        len(rows) >= settings.max_spreadsheet_rows
                        or total_cells + len(bounded_row) > settings.max_spreadsheet_cells
                        or total_chars + row_chars > settings.max_document_extract_chars
                    ):
                        truncated = True
                        break
                    rows.append(bounded_row)
                    total_cells += len(bounded_row)
                    total_chars += row_chars

            text_lines = [" | ".join(row) for row in rows[:100]]
            suffix = "\n[Data dipotong karena batas keamanan parser.]" if truncated else ""
            summary_text = (
                f"=== File CSV (Rows loaded: {len(rows)}) ===\n"
                + "\n".join(text_lines)
                + suffix
            )
            return summary_text, {"default": rows}
        except Exception as exc:
            return f"Error parsing CSV: {exc}", None


spreadsheet_parser = SpreadsheetParser()
